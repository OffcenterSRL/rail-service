from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import load_workbook


ROOT = Path("/Users/leonardo/Dev/rail-service")
PDF_DIR = ROOT / "data" / "pdf"
XLSX_PATH = ROOT / "data" / "excel" / "catalogo_completo.xlsx"
JSON_PATH = ROOT / "packages" / "frontend" / "src" / "assets" / "materials" / "materials.json"


@dataclass
class MaterialRow:
    tavola: str
    codice_fs: str
    descrizione: str
    riferimento: str
    codice_fornitore: str
    codice_ditta_documento: str
    numero_documento: str
    segno_matr_pos: str
    numero_pezzi: str
    unita_conto: str
    note: str
    source_pdf: str
    source_page: str

    def as_json(self) -> dict[str, str]:
        return {
            "tavola": self.tavola,
            "codice_fs": self.codice_fs,
            "descrizione": self.descrizione,
            "riferimento": self.riferimento,
            "codice_fornitore": self.codice_fornitore,
            "codice_ditta_documento": self.codice_ditta_documento,
            "numero_documento": self.numero_documento,
            "segno_matr_pos": self.segno_matr_pos,
            "numero_pezzi": self.numero_pezzi,
            "unita_conto": self.unita_conto,
            "note": self.note,
            "source_pdf": self.source_pdf,
            "source_page": self.source_page,
        }

    def as_xlsx_row(self) -> list[str]:
        return [
            self.tavola,
            self.codice_fs,
            self.descrizione,
            self.riferimento,
            self.codice_fornitore,
            self.codice_ditta_documento,
            self.numero_documento,
            self.segno_matr_pos,
            self.numero_pezzi,
            self.unita_conto,
            self.note,
        ]

    def dedupe_key(self) -> tuple[str, ...]:
        return (
            self.tavola,
            self.codice_fs,
            self.descrizione,
            self.riferimento,
            self.codice_fornitore,
            self.codice_ditta_documento,
            self.numero_documento,
            self.segno_matr_pos,
            self.numero_pezzi,
            self.unita_conto,
            self.note,
            self.source_pdf,
            self.source_page,
        )


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def merge_positions(values: Iterable[float], tolerance: float = 2.5) -> list[float]:
    merged: list[list[float]] = []
    for value in sorted(values):
        if not merged or abs(merged[-1][-1] - value) > tolerance:
            merged.append([value])
        else:
            merged[-1].append(value)
    return [sum(group) / len(group) for group in merged]


def group_words_by_line(words: list[dict], tolerance: float = 2.0) -> list[tuple[float, list[dict]]]:
    lines: list[tuple[float, list[dict]]] = []
    for word in sorted(words, key=lambda item: (item["top"], item["x0"])):
        top = float(word["top"])
        if not lines or abs(lines[-1][0] - top) > tolerance:
            lines.append((top, [word]))
        else:
            lines[-1][1].append(word)
    return lines


def line_text(words: list[dict]) -> str:
    return normalize_spaces(" ".join(word["text"] for word in sorted(words, key=lambda item: item["x0"])))


def extract_table_boundaries(page: pdfplumber.page.Page) -> list[float]:
    verticals: list[float] = []

    for rect in page.rects:
        width = abs(rect["x1"] - rect["x0"])
        height = abs(rect["y1"] - rect["y0"])
        if width <= 2 and height >= 120:
            verticals.extend([rect["x0"], rect["x1"]])

    for line in page.lines:
        if abs(line["x0"] - line["x1"]) <= 1 and abs(line["y1"] - line["y0"]) >= 120:
            verticals.extend([line["x0"], line["x1"]])

    return merge_positions(verticals)


def find_letter_line(lines: list[tuple[float, list[dict]]]) -> tuple[float, list[dict]] | None:
    for top, words in lines:
        if line_text(words).lower() == "a b c d e f g h i j":
            return top, words
    return None


def is_material_table_page(page: pdfplumber.page.Page) -> bool:
    text = page.extract_text() or ""
    if "Documento identificativo" not in text:
        return False
    lines = group_words_by_line(page.extract_words())
    if find_letter_line(lines) is None:
        return False
    return len(extract_table_boundaries(page)) >= 11


def extract_tavola(page: pdfplumber.page.Page) -> str:
    text = page.extract_text() or ""
    match = re.search(r"\bTav\.?\s+([A-Z0-9.]+)", text, re.IGNORECASE)
    return match.group(1).strip() if match else ""


def assign_words_to_columns(words: list[dict], boundaries: list[float]) -> list[str]:
    columns: list[list[dict]] = [[] for _ in range(10)]
    for word in words:
        center = (word["x0"] + word["x1"]) / 2
        for idx in range(10):
            if boundaries[idx] <= center < boundaries[idx + 1]:
                columns[idx].append(word)
                break
        else:
            if center >= boundaries[-1]:
                columns[-1].append(word)
    return [line_text(column_words) if column_words else "" for column_words in columns]


def merge_hyphenated(base: str, extra: str) -> str:
    if not extra:
        return base
    if not base:
        return extra
    if base.endswith("-"):
        return base[:-1] + extra
    return f"{base} {extra}".strip()


def is_footer_or_header(raw_line: str) -> bool:
    return bool(
        re.search(r"Catalogo n\.|Edizione|Pag\.|Documento identificativo|DESCRIZIONE|Riferim\.|NOTE|FS tavola", raw_line)
        and not re.search(r"\d", raw_line)
    )


def is_empty_row(columns: list[str]) -> bool:
    return not any(columns)


def is_continuation_row(columns: list[str]) -> bool:
    desc_only = columns[1] and not any(columns[idx] for idx in (0, 2, 3, 4, 5, 6, 7))
    note_only = columns[9] and not any(columns[idx] for idx in range(9))
    return desc_only or note_only


def row_has_material_payload(columns: list[str]) -> bool:
    return bool(columns[1] and any(columns[idx] for idx in (2, 3, 4, 5, 7, 9)))


def fix_units_and_notes(columns: list[str]) -> list[str]:
    if columns[8] and (not re.search(r"\d", columns[8]) or columns[8].lower().startswith("vedi")):
        columns[9] = normalize_spaces(f"{columns[8]} {columns[9]}")
        columns[8] = ""
    return columns


def extract_rows_from_page(pdf_name: str, page: pdfplumber.page.Page) -> list[MaterialRow]:
    boundaries = extract_table_boundaries(page)
    if len(boundaries) < 11:
        return []

    lines = group_words_by_line(page.extract_words())
    letter_line = find_letter_line(lines)
    if letter_line is None:
        return []

    letter_top = letter_line[0]
    tavola = extract_tavola(page)
    rows: list[MaterialRow] = []
    previous: MaterialRow | None = None

    for top, words in lines:
        if top <= letter_top:
            continue

        raw = line_text(words)
        if not raw or is_footer_or_header(raw):
            continue

        columns = assign_words_to_columns(words, boundaries[:11])
        if is_empty_row(columns):
            continue

        if is_continuation_row(columns) and previous is not None:
            if columns[1]:
                previous.descrizione = merge_hyphenated(previous.descrizione, columns[1])
            if columns[9]:
                previous.note = merge_hyphenated(previous.note, columns[9])
            continue

        columns = fix_units_and_notes(columns)
        if not row_has_material_payload(columns):
            continue

        row = MaterialRow(
            tavola=tavola,
            codice_fs=columns[0],
            descrizione=columns[1],
            riferimento=columns[2],
            codice_fornitore=columns[3],
            codice_ditta_documento=columns[4],
            numero_documento=columns[5],
            segno_matr_pos=columns[6],
            numero_pezzi=columns[7],
            unita_conto=columns[8],
            note=columns[9],
            source_pdf=pdf_name,
            source_page=str(page.page_number),
        )
        rows.append(row)
        previous = row

    return rows


def extract_all_rows() -> list[MaterialRow]:
    rows: list[MaterialRow] = []
    for pdf_path in sorted(PDF_DIR.glob("*.pdf")):
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                if not is_material_table_page(page):
                    continue
                rows.extend(extract_rows_from_page(pdf_path.name, page))

    unique: list[MaterialRow] = []
    seen: set[tuple[str, ...]] = set()
    for row in rows:
        key = row.dedupe_key()
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def write_json(rows: list[MaterialRow]) -> None:
    JSON_PATH.write_text(json.dumps([row.as_json() for row in rows], ensure_ascii=False, separators=(",", ":")))


def write_xlsx(rows: list[MaterialRow]) -> None:
    workbook = load_workbook(XLSX_PATH)
    worksheet = workbook[workbook.sheetnames[0]]

    if worksheet.max_row >= 3:
        worksheet.delete_rows(3, worksheet.max_row - 2)

    for row_idx, row in enumerate(rows, start=3):
        for col_idx, value in enumerate(row.as_xlsx_row(), start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    workbook.save(XLSX_PATH)


def main() -> None:
    rows = extract_all_rows()
    write_json(rows)
    write_xlsx(rows)
    print(f"Extracted {len(rows)} material rows")


if __name__ == "__main__":
    main()
